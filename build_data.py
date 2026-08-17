#!/usr/bin/env python3
"""
Rebuild the dashboard data from the two source Excel files and inject it into index.html.

Usage:
    python build_data.py \
        --framework Q2_FGDs.xlsx \
        --raw Focus_Group_Discussion-<export>.xlsx

Both the framework workbook (analysed findings) and the raw Kobo export are read,
normalised, joined, and written to data/data.json. The same JSON is embedded inline
into index.html between the DATA anchors, so the dashboard stays a single
self-contained file that works on GitHub Pages and by double-click.

Re-run this each quarter with the new exports to refresh the dashboard.
"""
import openpyxl, json, re, argparse, os, math

# ---- admin-1 map geometry (TopoJSON -> simplified GeoJSON) --------------
def _rdp(points, eps):
    n = len(points)
    if n < 3:
        return points
    keep = [False]*n; keep[0] = keep[-1] = True
    stack = [(0, n-1)]
    while stack:
        s, e = stack.pop()
        x1, y1 = points[s]; x2, y2 = points[e]
        dx, dy = x2-x1, y2-y1; d2 = dx*dx + dy*dy
        idx, dmax = -1, 0.0
        for i in range(s+1, e):
            x0, y0 = points[i]
            if d2 == 0:
                dist = ((x0-x1)**2 + (y0-y1)**2) ** 0.5
            else:
                tt = ((x0-x1)*dx + (y0-y1)*dy) / d2
                px, py = x1+tt*dx, y1+tt*dy
                dist = ((x0-px)**2 + (y0-py)**2) ** 0.5
            if dist > dmax:
                dmax, idx = dist, i
        if dmax > eps and idx != -1:
            keep[idx] = True; stack.append((s, idx)); stack.append((idx, e))
    return [p for p, k in zip(points, keep) if k]

def build_geo(topo_path, ndigits=4, eps=0.001):
    t = json.load(open(topo_path))
    scale, trans = t["transform"]["scale"], t["transform"]["translate"]
    arcs = []
    for arc in t["arcs"]:
        x = y = 0; pts = []
        for dx, dy in arc:
            x += dx; y += dy
            pts.append([x*scale[0]+trans[0], y*scale[1]+trans[1]])
        arcs.append(pts)
    def arc_coords(i):
        return list(reversed(arcs[~i])) if i < 0 else arcs[i]
    def ring_coords(ring):
        coords = []
        for i in ring:
            seg = arc_coords(i)
            coords.extend(seg[1:] if coords else seg)
        return coords
    def finish(ring):
        r = _rdp(ring, eps)
        if r[0] != r[-1]:
            r.append(r[0])
        if len(r) < 4:
            r = ring
        return [[round(x, ndigits), round(y, ndigits)] for x, y in r]
    features = []
    for gm in t["objects"]["MDA_Adm1_poly_GoM"]["geometries"]:
        name = gm["properties"]["ADM1_EN"]
        if gm["type"] == "Polygon":
            geom = {"type": "Polygon", "coordinates": [finish(ring_coords(r)) for r in gm["arcs"]]}
        else:
            geom = {"type": "MultiPolygon",
                    "coordinates": [[finish(ring_coords(r)) for r in poly] for poly in gm["arcs"]]}
        features.append({"type": "Feature",
                         "properties": {"name": name, "key": name.lower()},
                         "geometry": geom})
    return {"type": "FeatureCollection", "features": features}

# ---- name normalisation -------------------------------------------------
def norm_area(a):
    a = (a or "").strip()
    m = {
        "Legal Status": "Legal Status & Documentation",
        "Legal Status and Documentation": "Legal Status & Documentation",
        "Livelihoods": "Livelihoods & Basic Needs",
        "Livelihoods and basic needs": "Livelihoods & Basic Needs",
    }
    return m.get(a, a)

def clean_group(g):
    if not g:
        return "Unspecified"
    g = str(g)
    mapping = {
        "Older persons": "Older Persons (60+)",
        "Persons with dis": "Persons with Disabilities",
        "LGBTIQ": "LGBTIQ+",
        "Children and Tee": "Children & Teens",
        "People of pre-re": "Pre-Retirement (50\u201359)",
        "Adults (30-49)": "Adults (30\u201349)",
        "Ethnic Roma": "Ethnic Roma",
    }
    for k, v in mapping.items():
        if g.startswith(k) or k in g:
            return v
    return g.strip()

# ---- extraction ---------------------------------------------------------
def _split_recs(s):
    s = (s or "").strip()
    return [x.strip(" \u2022\t") for x in re.split(r"\n?\u2022|\n", s) if x.strip(" \u2022\t")]

def build(framework_path, raw_path, geo_path=None):
    # framework \u2014 multilingual (English / Ukrainian / Romanian).
    # Preferred source is the "Combined_Multilingual" sheet whose columns hold the
    # three languages row-for-row; findings and recommendations become {en,ro,uk}
    # objects, and per-language profile/area labels are collected into an i18n map.
    # Falls back to a single-language sheet (legacy 5-column layout) when the
    # multilingual sheet is absent, emitting the same {en,ro,uk} shape with the
    # other languages left blank so the dashboard can fall back to English.
    wb = openpyxl.load_workbook(framework_path, data_only=True)
    framework = []
    prof_i18n, area_i18n = {}, {}
    if "Combined_Multilingual" in wb.sheetnames:
        ws = wb["Combined_Multilingual"]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
        cl = lambda v: (v or "").strip()
        for r in rows:
            # cols: 0 prof(EN) 1 (UA) 2 (RO) | 3 area(EN) 4 (UA) 5 (RO)
            #       6 finding(EN) 7 (UA) 8 (RO) | 9 recs(EN) 10 (UA) 11 (RO)
            pEN, pUA, pRO = cl(r[0]), cl(r[1]), cl(r[2])
            aEN, aUA, aRO = cl(r[3]), cl(r[4]), cl(r[5])
            fEN, fUA, fRO = cl(r[6]), cl(r[7]), cl(r[8])
            recEN, recUA, recRO = _split_recs(r[9]), _split_recs(r[10]), _split_recs(r[11])
            area = norm_area(aEN)
            framework.append({
                "id": len(framework), "quarter": "Q2", "profile": pEN,
                "area": area, "area_original": aEN,
                "finding": {"en": fEN, "ro": fRO, "uk": fUA},
                "recommendations": [{"en": e, "ro": o, "uk": u}
                                    for e, o, u in zip(recEN, recRO, recUA)],
            })
            prof_i18n.setdefault(pEN, {"en": pEN, "ro": pRO, "uk": pUA})
            area_i18n.setdefault(area, {"en": area, "ro": aRO, "uk": aUA})
    else:
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[1:]:
            if not any(r):
                continue
            area = norm_area(r[2])
            prof = (r[1] or "").strip()
            framework.append({
                "id": len(framework), "quarter": r[0], "profile": prof,
                "area": area, "area_original": (r[2] or "").strip(),
                "finding": {"en": (r[3] or "").strip(), "ro": "", "uk": ""},
                "recommendations": [{"en": x, "ro": "", "uk": ""} for x in _split_recs(r[4])],
            })
            prof_i18n.setdefault(prof, {"en": prof, "ro": prof, "uk": prof})
            area_i18n.setdefault(area, {"en": area, "ro": area, "uk": area})

    # raw coverage
    wb2 = openpyxl.load_workbook(raw_path, data_only=True)
    ws2 = wb2["Focus Group Discussion 2026"]
    d = list(ws2.iter_rows(values_only=True))
    hdr = d[0]
    def ci(name):
        for i, h in enumerate(hdr):
            if h and str(h).strip() == name:
                return i
        return None
    c_id = ci("_id")
    age_cols = list(range(39, 48))
    age_labels = ["8-10", "11-14", "15-17", "18-29", "30-49", "50-59", "60-69", "70-79", "80+"]

    # per-topic ("Number of persons answering") counts, by questionnaire section,
    # folded to the framework sectors used across the dashboard. Sectors backed by
    # several questionnaire sections (Livelihoods & Basic Needs) take the largest
    # sub-section per FGD, so a topic count never exceeds the group size.
    SECTION_TO_SECTOR = {
        "Legal Status": "Legal Status & Documentation",
        "Integration and Inclusion": "Integration & Inclusion",
        "Housing": "Housing",
        "Access to medical assistance:": "Healthcare",
        "Access to Social Services and Assistance:": "Livelihoods & Basic Needs",
        "Access to basic needs": "Livelihoods & Basic Needs",
        "Education": "Education",
        "Employment and livelihoods": "Livelihoods & Basic Needs",
        "Access to information": "Access to Information",
        "Peaceful co-existence and community cohesion": "Peaceful Coexistence",
    }
    SECTION_COLS = {p: (ci(p + "/Number of persons answering:"), ci(p + "/Males:"),
                        ci(p + "/Females:"), ci(p + "/Other Gender:")) for p in SECTION_TO_SECTOR}
    def numv(row, idx):
        v = row[idx] if idx is not None else None
        return v if isinstance(v, (int, float)) and not isinstance(v, bool) else 0
    def fgd_topics(row):
        # session-level gender shares, used to apportion topics whose own
        # gender breakdown was left blank (e.g. the Bender session)
        sm, sf, so = numv(row, 20), numv(row, 21), numv(row, 22)
        stot = sm + sf + so
        out = {}
        for sect, sector in SECTION_TO_SECTOR.items():
            ni, mi, fi, oi = SECTION_COLS[sect]
            n = numv(row, ni)
            if n <= 0:
                continue
            m, f, o = numv(row, mi), numv(row, fi), numv(row, oi)
            if m + f + o == 0 and stot > 0:
                # largest-remainder apportionment of n by the intro split
                exact = [n*sm/stot, n*sf/stot, n*so/stot]
                base = [int(v) for v in exact]
                for _ in range(n - sum(base)):
                    i = max(range(3), key=lambda j: exact[j]-base[j])
                    base[i] += 1
                m, f, o = base
            cur = {"n": n, "m": m, "f": f, "o": o}
            if sector not in out or cur["n"] > out[sector]["n"]:
                out[sector] = cur
        return out

    coverage, age_totals = [], [0]*9
    g_tot = {"male": 0, "female": 0, "other": 0}
    for r in d[1:]:
        if not r[c_id]:
            continue
        ages = [(r[a] or 0) if isinstance(r[a], (int, float)) else 0 for a in age_cols]
        for j, v in enumerate(ages):
            age_totals[j] += v
        m, f, o = r[20] or 0, r[21] or 0, r[22] or 0
        g_tot["male"] += m if isinstance(m, (int, float)) else 0
        g_tot["female"] += f if isinstance(f, (int, float)) else 0
        g_tot["other"] += o if isinstance(o, (int, float)) else 0
        coverage.append({
            "id": r[c_id], "raion": str(r[11]).strip() if r[11] else "",
            "location": str(r[12]).strip() if r[12] else "", "group": clean_group(r[25]),
            "org": str(r[13]).strip() if r[13] else "", "date": str(r[10])[:10] if r[10] else "",
            "n": r[19] or 0, "male": m, "female": f, "other": o, "ages": ages,
            "topics": fgd_topics(r),
        })

    # quotes
    wsq = wb2["quotes_group"]
    q = list(wsq.iter_rows(values_only=True))
    id_to_group = {r["id"]: r["group"] for r in coverage}
    id_to_loc = {r["id"]: r["location"] for r in coverage}
    quotes = []
    for i, r in enumerate(q[1:]):
        txt = r[0]
        if not txt or not str(txt).strip():
            continue
        txt = str(txt).strip().strip('"').strip(",").strip()
        sid = r[4]
        quotes.append({"id": len(quotes), "text": txt, "group": id_to_group.get(sid, ""),
                       "location": id_to_loc.get(sid, ""), "date": str(r[6])[:10] if r[6] else ""})

    from collections import Counter
    by_group, part_by_group = Counter(), Counter()
    by_raion, part_by_raion = Counter(), Counter()
    for r in coverage:
        by_group[r["group"]] += 1; part_by_group[r["group"]] += r["n"]
        by_raion[r["raion"]] += 1; part_by_raion[r["raion"]] += r["n"]

    geo = build_geo(geo_path) if geo_path and os.path.exists(geo_path) else None

    dates = [r["date"] for r in coverage if r["date"]]
    return {
        "geo": geo,
        "meta": {"quarter": framework[0]["quarter"] and (str(framework[0]["quarter"]) + " 2026") or "Q2 2026",
                 "n_fgds": len(coverage), "n_participants": sum(r["n"] for r in coverage),
                 "gender": g_tot, "n_raions": len(set(r["raion"] for r in coverage)),
                 "date_range": [min(dates), max(dates)] if dates else ["", ""]},
        "age_labels": age_labels, "age_totals": age_totals,
        "coverage": coverage, "framework": framework, "quotes": quotes,
        "i18n": {"profile": prof_i18n, "area": area_i18n},
        "agg": {"fgds_by_group": dict(by_group), "part_by_group": dict(part_by_group),
                "fgds_by_raion": dict(by_raion), "part_by_raion": dict(part_by_raion)},
    }

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--framework", default="Q2_FGDs.xlsx")
    ap.add_argument("--raw", default="Focus_Group_Discussion-2026-07-30-13-33-50.xlsx")
    ap.add_argument("--geo", default="MDA_Adm1_poly_GoM__Transnitrian_Region_.json")
    args = ap.parse_args()

    data = build(args.framework, args.raw, args.geo)
    os.makedirs("data", exist_ok=True)
    with open("data/data.json", "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, separators=(",", ":"))
    print(f"OK — {data['meta']['n_fgds']} FGDs, {data['meta']['n_participants']} participants, "
          f"{len(data['framework'])} findings, {len(data['quotes'])} quotes.")
    print("Wrote data/data.json. index.html reads this file at runtime — no rebuild of the HTML needed.")
